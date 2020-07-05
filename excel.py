from main import *
from openpyxl import Workbook, load_workbook


def main():
    """
    This program generate the excel of results.
    """

    # open the excel template
    wb = load_workbook("excel_output_template.xlsx")

    # read the dataset
    print("start reading dataset...")
    extensions = ["jpg", "jpeg"]
    data_set = read_data_set("dataset/", extensions)
    print("data set successfully read")

    # read the 3 sheets in the excel file
    feature_extraction = wb["Feature Extraction"]
    similarity_measure = wb["Similarity Measure"]
    p_r = wb["P-R"]

    # some offset to move in cells
    offset1 = 12  # the table of feature extraction start at the 12th cell vertically
    offset2 = 4  # the table of similarity measure start at the 4th cell vertically
    offset3 = 14  # each query has 14 cells between them vertically
    p_r_offset_start = 7  # the first query table start at the 7th cell vertically
    p_r_increment = 14  # used to increment trough the query in the p/r curve sheet

    # read the images of the dataset
    index = 0
    for image in data_set:
        # get the feature vector
        v = image.get_feature_vector()

        # compute the top 10 similarity for this image and the others from the data set with 3 digits
        top_n = top_n_similarity(10, generate_euclidean_set(data_set, image, 3))

        # write the top 10 in the p/r curve sheet
        index_top_n = 0
        for el in top_n:
            # write the ID of the image which is the name of the jpeg image
            p_r["C" + str(p_r_offset_start+index_top_n)].value = get_image_by_id(data_set, el[0]).full_path.split("/")[2].split(".")[0]
            index_top_n += 1

        p_r_offset_start += p_r_increment

        # write the feature vector in the first table
        feature_extraction['C' + str(index + offset1)].value = v["mean"]
        feature_extraction['D' + str(index + offset1)].value = v["variance"]
        feature_extraction['E' + str(index + offset1)].value = v["skewness"]
        feature_extraction['F' + str(index + offset1)].value = v["contrast"]
        feature_extraction['G' + str(index + offset1)].value = v["dissimilarity"]
        feature_extraction['H' + str(index + offset1)].value = v["homogeneity"]
        feature_extraction['I' + str(index + offset1)].value = v["ASM"]
        feature_extraction['J' + str(index + offset1)].value = v["energy"]
        feature_extraction['K' + str(index + offset1)].value = v["correlation"]

        # write the feature vector in the second table
        similarity_measure['C' + str(index + offset2)].value = v["mean"]
        similarity_measure['D' + str(index + offset2)].value = v["variance"]
        similarity_measure['E' + str(index + offset2)].value = v["skewness"]
        similarity_measure['F' + str(index + offset2)].value = v["contrast"]
        similarity_measure['G' + str(index + offset2)].value = v["dissimilarity"]
        similarity_measure['H' + str(index + offset2)].value = v["homogeneity"]
        similarity_measure['I' + str(index + offset2)].value = v["ASM"]
        similarity_measure['J' + str(index + offset2)].value = v["energy"]
        similarity_measure['K' + str(index + offset2)].value = v["correlation"]

        # generate the matrix of euclidean distances
        j = 0
        for img in data_set:
            similarity_measure.cell(row=index+offset2, column=j+offset3).value = euclidean_distance(img, image, 3)
            j += 1

        index += 1

    # save the new excel file
    wb.save("out.xlsx")


if __name__ == '__main__':
    main()